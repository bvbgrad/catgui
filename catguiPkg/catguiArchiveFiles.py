import datetime
import hashlib
import logging
import os

from openpyxl import Workbook, load_workbook
from pathlib import Path
from shutil import copy2, disk_usage

import PySimpleGUI as sg

import catgui

from catgui import log_wrap

LOGGER_NAME = "catgui"
logger = logging.getLogger(LOGGER_NAME)


@log_wrap
def create_backup(scan_file_name):
    logger.info(f"create_backup({scan_file_name})")
    args = catgui.getargs()

    wb = load_workbook(filename=scan_file_name, read_only=True, data_only=True)

    start_folder = ""
    target_folder = ""
    save_size = 0


    ws_summary = wb["Summary"]
    for value in ws_summary.iter_rows(min_col=1, max_col=2, values_only=True):
        if args.verbose: print(value)
        if value[0] == "start_folder":
            start_folder = value[1]
            start_path = Path(start_folder)
            print(f"start_path {start_path}")
        if value[0] == "target_folder":
            target_folder = value[1]
            target_path = Path(target_folder)
            print(f"target_path {target_path}")
        if value[0] == "total save size":
            save_size = int(value[1])

    if os.path.isdir(target_path):
        total, used, free_space = disk_usage(target_path)
    else:
        return ("No target folder")

    if free_space < save_size:
        print(f"Required disk space: {save_size},  Available space: {free_space}")
        return("Not enough room on disk to for backup")

    files_copied = 0
    file_collisions = 0
    files_processed = 0
    files_skipped = 0

    ws_files = wb["Files"]
    for value in ws_files.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True):
        source_file_path = Path(Path(value[0]) / value[1])

        destination_file_folder = value[0].replace(start_folder, target_folder)
        destination_folder_path = Path(destination_file_folder)
        try:
            if not os.path.exists(destination_folder_path):
                os.makedirs(destination_folder_path, exist_ok=True)
            destination_file_path = Path(destination_folder_path / value[1])

            files_processed += 1
            if value[5] == "hash":
                if os.path.exists(destination_file_path):
                    files_skipped += 1
                    continue # skip this file and process the next one

                copy2(source_file_path, destination_file_path)
                if args.verbose: print(f"Copy {source_file_path} to \n\t{destination_file_path}")
                files_copied += 1
            else:
                file_collisions += 1
        # except OSError:
        #     print("No more space on the target disk")
        #     logger.info(f"No more space on the target disk")
        except IsADirectoryError:
            print("Destination File is Directory")
            logger.info("Destination File is Directory")
        except IOError:
            print("Input Output operation get Failed")
            logger.info("Input Output operation get Failed")
        except PermissionError:
            print("Don't have the Permission to Copy a file")
            logger.info("Don't have the Permission to Copy a file")
        except Exception as e:
            print(f"Unexpected error {e} occured")
            logger.info(f"Unexpected error {e} occured")

    print(f"Files processed: {files_processed}  Files skipped: {files_skipped}")
    print(f"Number of additional duplicates: {file_collisions}")
    if args.verbose: print(f"Number of additional duplicates: {file_collisions}")
    return (f"Copied {files_copied} of {files_processed} files")


@log_wrap
def scan_files(scan_parameters):
    logger.info(f"archive_scan_files()")

    args = catgui.getargs()
    if args.verbose: print(f"Options: {args}")

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'Summary'

    ws1['A1']= f"Scan summary "
    for parameter_key in sorted(scan_parameters):
        if isinstance(scan_parameters[parameter_key], list):
            for item in scan_parameters[parameter_key]:
                ws1.append((parameter_key, item))
        else:
            ws1.append((parameter_key, scan_parameters[parameter_key]))

    summary = save2(args, wb, scan_parameters)

    for summary_key in summary:
        ws1.append((summary_key, summary[summary_key]))

    filename = f'backup_scan_{datetime.datetime.now().strftime("%Y%m%dt%H%M%S")}.xlsx'
    if 'target_folder' in scan_parameters:
        filepath = Path(scan_parameters['target_folder'], filename)
        wb.save(filepath)
        logger.info(f"Scan summary saved: {filepath}")

    wb.close()


@log_wrap
def file_survey(scan_parameters):
    logger.info("file_survey()")

    filelist = []
    ignorelist = []
    start_path = scan_parameters['start_folder']
    for parent, dirs, files in os.walk(start_path):
        for filename in files: #row[2] is a tuple of the filenames
            full_path: Path = Path(parent / Path(filename)) # row[0] is the parent directory
            st = os.stat(full_path)
            filelist.append([
                parent, filename, st.st_size, 
                datetime.datetime.fromtimestamp(st.st_ctime), 
                datetime.datetime.fromtimestamp(st.st_mtime),
                'hash'
            ])
    
    print(f"number of files scanned = {len(filelist)}")
    excluded_count = 0
    keeplist = []
    for file_object in filelist:
        _exclude = False
        for item in scan_parameters["exclude_list"]:
            if item in file_object[0]:
                _exclude = True
        if _exclude:
            excluded_count += 1
            ignorelist.append(file_object)
        else:
            keeplist.append(file_object)
        
    print(f"number of ignored entries {len(ignorelist)}")
    print(f"number of entries retained {len(keeplist)}")

    return (keeplist, ignorelist)


@log_wrap
def save2(args, wb, scan_parameters):
    logger.info("save2()")

    header_row = ('Directory', 'Filename', 'size', 'created', 'modified', 'blake2')

    ws2 = wb.create_sheet('Files')
    ws2.append(header_row)
    ws2.auto_filter.ref = "A:F"
    ws2.freeze_panes = "B2"

    ws2d = wb.create_sheet('Dups')
    ws2d.append(header_row)
    ws2d.auto_filter.ref = "A:F"
    ws2d.freeze_panes = "B2"

    ws2i = wb.create_sheet('Ignore')
    ws2i.append(header_row)
    ws2i.auto_filter.ref = "A:F"
    ws2i.freeze_panes = "B2"

    keeplist, ignorelist = file_survey(scan_parameters)

    logger.info(f"Begin data save: keep candidates {len(keeplist)}")
    logger.info(f"Begin data save: ignore list files {len(ignorelist)}")
    for file_object in ignorelist: #save those that were ignored in the file survey
        ws2i.append(file_object)

    number_files = 0
    number_save = 0
    size_save = 0
    number_ignore = 0
    size_ignore = 0
    duplicate_count_list = {}
    size_duplicates = 0

    DATE_FORMAT = "%Y-%m-%d"
    scan_start_time = datetime.datetime.strptime(scan_parameters['scan_start_date'], DATE_FORMAT)

# add files to the 'keep' sheet if their mod date is newer than the scan start date
# When more than one 'copy' is found (based on filename concatenated with its file size),
# Compute the hash digest and add the file object to the 'dups' sheet
# This situation indicates a probable duplicate file (see profile section)

    for file_object in keeplist:
        number_files += 1
        if scan_start_time < file_object[4]:  #keep files with scan start before modified time
            number_save += 1
            size_save += file_object[2]

            filename_size = file_object[1] + str(file_object[2])
            if filename_size in duplicate_count_list.keys():
                duplicate_count_list[filename_size] += 1
                size_duplicates += file_object[2]

                full_path = Path(file_object[0]) / Path(file_object[1])
                try:
                    file_hash = hashlib.blake2b()
                    with open(full_path, "rb") as f:
                        while chunk := f.read(8192):
                            file_hash.update(chunk)
                    file_object[5] = file_hash.hexdigest()[:20]
                    ws2d.append(file_object)
                except PermissionError as e:
                    sg.popup_quick_message(f'exception {e}', 
                        'Check if a previous Excel scan file is still open.', 
                        keep_on_top=True, auto_close_duration=5,
                        background_color='yellow', text_color='black')
                    break

            else:  # put the 'new' filename_size on the list
                duplicate_count_list[filename_size] = 1

            ws2.append(file_object)  # save all those found within the scan period
        else:
            ws2i.append(file_object)  # add the rest to ignore tab
            number_ignore += 1
            size_ignore += file_object[2]

    print(f"Total files processed {number_files}")
    print(f"  Number of modified files {number_save}")
    print(f"  Number of files ignored {number_ignore}")

    duplicates = list(duplicate_count_list.values())
    duplicate_count = 0
    for duplicate in duplicates:
        if duplicate > 1:
            duplicate_count += 1
    print(f"  Number of candidate duplicates {duplicate_count}")

    summary = {
    'total files' : number_files,
    'Modified files after scan date' : number_save, 
    'total save size' : size_save,
    'files before scan date (ignored)' : number_ignore,
    'total ignore size' : size_ignore,
    'candidate duplicates' : duplicate_count,
    'duplicates size estimate' : size_duplicates
    }

    return summary
